from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
UCDB_FILE = (
    BASE_DIR
    / "data"
    / "external"
    / "ucdb"
    / "temp"
    / "GHS_UCDB_REGION_CENTRAL_AND_SOUTHERN_ASIA_R2024A.xlsx"
)
RAW_DIR = BASE_DIR / "data" / "raw"
OUTPUT_SCENARIOS = RAW_DIR / "landuse_scenarios.csv"
OUTPUT_METRICS = RAW_DIR / "ucdb_mumbai_metrics.csv"


def read_mumbai_ghsl_record(
    path: Path, mumbai_id: int = 7599
) -> dict[str, float | str]:
    if not path.exists():
        raise FileNotFoundError(f"UCDB workbook not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["GHSL"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)

        needed = [
            "ID_UC_G0",
            "GC_UCN_MAI_2025",
            "GC_UCA_KM2_2025",
            "GH_BUS_TOT_2000",
            "GH_BUS_TOT_2025",
        ]
        idx = {name: header.index(name) for name in needed}

        for row in rows:
            if row[idx["ID_UC_G0"]] == mumbai_id:
                return {
                    "id_uc_g0": int(row[idx["ID_UC_G0"]]),
                    "city_name": str(row[idx["GC_UCN_MAI_2025"]]),
                    "area_km2": float(row[idx["GC_UCA_KM2_2025"]]),
                    "built_surface_m2_2000": float(
                        row[idx["GH_BUS_TOT_2000"]]
                    ),
                    "built_surface_m2_2025": float(
                        row[idx["GH_BUS_TOT_2025"]]
                    ),
                }
    finally:
        workbook.close()

    raise ValueError(
        f"Mumbai record with ID_UC_G0={mumbai_id} not found in GHSL sheet"
    )


def build_landuse_scenarios(record: dict[str, float | str]) -> pd.DataFrame:
    area_km2 = float(record["area_km2"])
    area_m2 = area_km2 * 1_000_000

    imp_pre = float(record["built_surface_m2_2000"]) / area_m2
    imp_post = float(record["built_surface_m2_2025"]) / area_m2

    perv_pre = 1.0 - imp_pre
    perv_post = 1.0 - imp_post

    # Standard Rational Method coefficients for urbanized vs pervious surfaces.
    c_impervious = 0.90
    c_pervious = 0.30

    c_pre = imp_pre * c_impervious + perv_pre * c_pervious
    c_post = imp_post * c_impervious + perv_post * c_pervious

    scenarios = pd.DataFrame(
        [
            {
                "scenario": "pre_urban",
                "area_km2": area_km2,
                "impervious_fraction": imp_pre,
                "pervious_fraction": perv_pre,
                "runoff_coefficient_weighted": c_pre,
            },
            {
                "scenario": "post_urban",
                "area_km2": area_km2,
                "impervious_fraction": imp_post,
                "pervious_fraction": perv_post,
                "runoff_coefficient_weighted": c_post,
            },
        ]
    )
    return scenarios


def main() -> None:
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    record = read_mumbai_ghsl_record(UCDB_FILE)
    scenarios = build_landuse_scenarios(record)

    metrics = pd.DataFrame([record])
    metrics.to_csv(OUTPUT_METRICS, index=False)
    scenarios.to_csv(OUTPUT_SCENARIOS, index=False)

    print("Prepared land-use scenarios from UCDB GHSL for Mumbai")
    print(f"Saved metrics: {OUTPUT_METRICS}")
    print(f"Saved scenarios: {OUTPUT_SCENARIOS}")


if __name__ == "__main__":
    main()
